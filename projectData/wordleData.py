import numpy.random
from utilities.Constants import NUM_OF_GUESSES
import openpyxl

class wordleData:

    def __init__(self):
        self.book = openpyxl.load_workbook("C:\\Users\\omerh\\PycharmProjects\\"
                                      "nyt_games_testing_demo\\projectData\\wordle.xlsx")
        self.words =[]
        self.weights =[]
        occurrence = []
        sheet = self.book.active
        for i in range(2,sheet.max_row+1):
            self.words.append(sheet.cell(row=i,column=1).value.upper())
            occurrence.append(float(sheet.cell(row=i,column=2).value))
        sum_occurence = sum(occurrence)
        for j in range(len(occurrence)):
            self.weights.append(occurrence[j]/sum_occurence)


    def getWordleSet(self):
        return tuple(numpy.random.choice(self.words,NUM_OF_GUESSES,replace=False,p=self.weights))

    def getSets(self,number):
        sets =[]
        for i in range(number):
            sets.append(self.getWordleSet())
        return sets
